import openpyxl

def extraer_tabla_excel(archivo_excel, nombre_hoja):
    # Cargar el libro de trabajo
    libro = openpyxl.load_workbook(archivo_excel)

    # Seleccionar la hoja de cálculo
    hoja = libro[nombre_hoja]

    # Obtener las dimensiones de la hoja
    filas = hoja.max_row
    columnas = hoja.max_column
    print(f' fila maxima : {filas}, columna maxima : {columnas}')

    # Crear una lista para almacenar los datos de la tabla
    tabla = []

    # Iterar sobre las filas y columnas para extraer los datos
    for fila in range(1, filas + 1):
        fila_datos = []
        for columna in range(1, columnas + 1):
            celda = hoja.cell(row=fila, column=columna)
            fila_datos.append(celda.value)
            if celda.value == None:
                pass
        tabla.append(fila_datos)

    # Cerrar el libro de trabajo
    libro.close()

    return tabla

# Ejemplo de uso
archivo_excel = "PEDIDITO_DB.xlsx"  # Reemplaza con el nombre de tu archivo Excel
nombre_hoja = "Semana 1 del 21 al 27 de agost"  # Reemplaza con el nombre de tu hoja de cálculo

tabla_extraida = extraer_tabla_excel(archivo_excel, nombre_hoja)
print(tabla_extraida)
